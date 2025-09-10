from openpyxl import load_workbook

# If it finds a copy of the Output excel the script will resume from where it left
try:
    wb = load_workbook("Output.xlsx")
    print("Resuming...")
except:
    wb = load_workbook("Input.xlsx")
    print("Starting from the beginning...")

ws = wb.active

first_row = 2; # Enter the number of the first row where you put coordinates inside the input Excel
mr = ws.max_row
save_every_x_rows = 5 # Enter how often do you want to save the file (recommended for large sets of coordinates)

# Looks for the last row saved
def lastDone():
    for row in range(first_row, mr + 1):
        if(ws["B" + str(row)].value == None and row <= mr):
            return (int(row) - 1)
    print("Reverse Geocoding Completed, exiting loop...")
    return None

# Loads the coordinates left to do
def loadCoordinates(lastRow):
    coordsArray = []
    lastRow += 1
    for row in range(lastRow, mr + 1):
        coordsArray.append(str(ws["A" + str(row)].value))
    return coordsArray

def coordsExcel(lastRow, address, number):
    ws["B" + str(lastRow)] = address
    if(number[0] == save_every_x_rows or lastRow == mr):
        saveExcel()
        number[0] = 0

def saveExcel():
    print("Saving...")
    wb.save("Output.xlsx")